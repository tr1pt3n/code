# script này sẽ xóa đi các cột A và B
import os
import shlex
import openpyxl
import re

# Load the Excel file
workbook = openpyxl.load_workbook('/Users/tr1pt3n/Developer/LOTL/scripts.xlsx')

# Get the active worksheet
worksheet = workbook.active

# Get the maximum row number
max_row = worksheet.max_row

# Iterate over each row
for row in range(2, max_row+1):
    # Get the command string from column B
    command_string = worksheet.cell(row=row, column=2).value

    # Tokenize the command string using shlex
    # tokens = shlex.split(command_string)
    tokens = re.split('[/\\\.\s]', command_string)
    tokens = [re.sub('[\"\']', '', token) for token in tokens if token != '']

    # Write the tokens back to column C
    token_string = str(tokens)
    worksheet.cell(row=row, column=3).value = token_string

# Delete columns A and B
worksheet.delete_cols(1, 2)

# Save the updated workbook
workbook.save('/Users/tr1pt3n/Developer/LOTL/commands_with_tokens.xlsx')


# script này giữ lại cột A và B
# import os
# import shlex
# import openpyxl
# import re

# # Load the Excel file
# workbook = openpyxl.load_workbook('/Users/tr1pt3n/Developer/LOTL/scripts.xlsx')

# # Get the active worksheet
# worksheet = workbook.active

# # Get the maximum row number
# max_row = worksheet.max_row

# # Iterate over each row
# for row in range(2, max_row+1):
#     # Get the command string from column B
#     command_string = worksheet.cell(row=row, column=2).value

#     # Tokenize the command string using shlex
#     # tokens = shlex.split(command_string)
#     tokens = re.split('[/\\\.\s]', command_string)
#     tokens = [re.sub('[\"\']', '', token) for token in tokens if token != '']


#     # Write the tokens back to column C
#     token_string = str(tokens)
#     worksheet.cell(row=row, column=3).value = token_string

# # Save the updated workbook
# workbook.save('/Users/tr1pt3n/Developer/LOTL/commands_with_tokens.xlsx')